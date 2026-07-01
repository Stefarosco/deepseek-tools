"""
Excel 可视化工具 — 快速摸底 / HTML 预览 / 差异对比
========================================================
用途：让 AI 助手能"看见" Excel 的布局、颜色、边框、合并单元格等视觉效果
依赖：pip install openpyxl

用法：
  python excel_viz.py dump  <文件.xlsx>            # 快速摸底：Sheet 结构预览
  python excel_viz.py html  <文件.xlsx> [输出.html] # 导出保留视觉样式的 HTML
  python excel_viz.py diff  <文件A.xlsx> <文件B.xlsx>  # 对比两个文件的差异

示例：
  python excel_viz.py dump  D:\表格\销售数据.xlsx
  python excel_viz.py html  D:\表格\销售数据.xlsx preview.html
  python excel_viz.py diff  D:\表格\v1.xlsx D:\表格\v2.xlsx
"""

import os
import sys
import argparse
import re
import json
from datetime import datetime, time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from copy import copy


# ────────────────────────────────────────
#  工具函数
# ────────────────────────────────────────

def col_letter(col_idx):
    """1 → 'A', 27 → 'AA'"""
    return get_column_letter(col_idx)


def safe_str(val):
    """安全地把单元格值转成字符串"""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    return str(val)


def color_to_hex(color_obj):
    """尝试把 openpyxl 的颜色转成十六进制字符串"""
    if color_obj is None:
        return None
    try:
        if color_obj.type == "rgb" and color_obj.rgb and len(str(color_obj.rgb)) == 8:
            return "#" + str(color_obj.rgb)[2:]
        if color_obj.type == "theme":
            return f"theme({color_obj.theme})"
        if color_obj.type == "indexed":
            return f"indexed({color_obj.index})"
    except Exception:
        pass
    # 用 str 回退
    raw = str(color_obj.rgb) if color_obj and color_obj.rgb else None
    if raw and raw != "00000000":
        return "#" + raw[2:] if len(raw) >= 8 else "#" + raw
    return None


def font_css(font):
    """从 openpyxl Font 对象生成 CSS 样式片段"""
    parts = []
    if font.bold:
        parts.append("font-weight:bold")
    if font.italic:
        parts.append("font-style:italic")
    if font.underline:
        parts.append("text-decoration:underline")
    if font.size:
        parts.append(f"font-size:{font.size}pt")
    fg = color_to_hex(font.color)
    if fg and fg != "#000000":
        parts.append(f"color:{fg}")
    if font.name:
        parts.append(f"font-family:'{font.name}'")
    return ";".join(parts)


def fill_css(fill):
    """从 openpyxl PatternFill 生成背景色 CSS"""
    if fill and fill.patternType and fill.fgColor:
        c = color_to_hex(fill.fgColor)
        if c and c != "#000000" and c != "#00000000":
            return f"background-color:{c}"
    return ""


def border_css(cell_border):
    """从 openpyxl Border 生成边框 CSS"""
    styles = {
        "thin": "1px solid #999",
        "medium": "2px solid #666",
        "thick": "3px solid #333",
        "hair": "0.5px solid #ccc",
        "dashed": "1px dashed #999",
        "dotted": "1px dotted #999",
        "double": "3px double #666",
        "dashDot": "1px dashed #999",
        "mediumDashDot": "2px dashed #666",
    }

    def side_style(side_obj):
        if side_obj is None:
            return None
        # Map openpyxl's style string
        style_name = str(side_obj.style) if side_obj.style else None
        return styles.get(style_name, "1px solid #999" if style_name else None)

    parts = []
    if cell_border:
        for direction, attr in [("top", "top"), ("right", "right"),
                                  ("bottom", "bottom"), ("left", "left")]:
            s = side_style(getattr(cell_border, attr, None))
            if s:
                parts.append(f"border-{direction}:{s}")
    return ";".join(parts)


def align_css(alignment):
    """从 openpyxl Alignment 生成 CSS"""
    parts = []
    if alignment:
        h_map = {"center": "center", "left": "left", "right": "right",
                 "justify": "justify", "distributed": "justify"}
        v_map = {"center": "middle", "top": "top", "bottom": "bottom"}
        if alignment.horizontal in h_map:
            parts.append(f"text-align:{h_map[alignment.horizontal]}")
        if alignment.vertical in v_map:
            parts.append(f"vertical-align:{v_map[alignment.vertical]}")
        if alignment.wrap_text:
            parts.append("white-space:pre-wrap;word-wrap:break-word")
    return ";".join(parts)


# ────────────────────────────────────────
#  Dump 命令：快速摸底
# ────────────────────────────────────────

def cmd_dump(filepath):
    """打印 Excel 文件的所有 Sheet 结构信息"""
    wb = load_workbook(filepath, data_only=False)
    print(f"📊 文件: {filepath}")
    print(f"📋 Sheet 数量: {len(wb.sheetnames)}")
    print("=" * 70)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        dim = ws.dimensions
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        merged = ws.merged_cells.ranges

        print(f"\n── Sheet: 「{sheet_name}」──")
        print(f"   尺寸: {rows} 行 × {cols} 列  (范围: {dim})")

        # 合并单元格
        if merged:
            print(f"   合并单元格 ({len(merged)} 处):")
            for m in sorted(merged, key=str):
                print(f"     · {m}")
        else:
            print(f"   合并单元格: 无")

        # 前 5 行预览（只显示有数据的行列）
        print(f"   数据预览 (前 5 行):")
        preview_rows = min(5, rows)
        max_col = min(cols, 15)  # 最多显示 15 列
        for r in range(1, preview_rows + 1):
            vals = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                v = safe_str(cell.value)
                # 截断过长文本
                if len(v) > 40:
                    v = v[:37] + "..."
                vals.append(v)
            row_str = " | ".join(vals)
            if row_str.strip().replace("|", "").strip():
                print(f"    第{r}行: {row_str}")

        if cols > 15:
            print(f"    ... (省略了 {cols - 15} 列)")

        # 公式单元格
        formulas = []
        for r in range(1, min(rows + 1, 1001)):  # 最多扫 1000 行
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{col_letter(c)}{r}: {cell.value}")
        if formulas:
            print(f"   公式单元格 ({len(formulas)} 个):")
            for f in formulas[:20]:
                print(f"     · {f}")
            if len(formulas) > 20:
                print(f"     ... 还有 {len(formulas) - 20} 个公式")
        else:
            print(f"   公式单元格: 无")

        # 数据有效性和筛选
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"   自动筛选: {ws.auto_filter.ref}")

    wb.close()
    print("\n" + "=" * 70)
    print("✅ 摸底完成")


# ────────────────────────────────────────
#  HTML 命令：导出视觉预览
# ────────────────────────────────────────

def cmd_html(filepath, output_path=None):
    """把 Excel 导出为保留视觉样式的 HTML 文件"""
    if output_path is None:
        base = os.path.splitext(os.path.basename(filepath))[0]
        output_path = f"{base}_preview.html"

    wb = load_workbook(filepath, data_only=True)

    # 收集每个 Sheet 合并单元格信息，用于 rowspan/colspan
    merge_map = {}  # sheet_name -> {(r,c): (rowspan,colspan)}
    for sn in wb.sheetnames:
        ws = wb[sn]
        merge_map[sn] = {}
        for merged_range in ws.merged_cells.ranges:
            min_r, min_c = merged_range.min_row, merged_range.min_col
            rowspan = merged_range.max_row - min_r + 1
            colspan = merged_range.max_col - min_c + 1
            merge_map[sn][(min_r, min_c)] = (rowspan, colspan)
            # 标记被合并覆盖的单元格（在 HTML 渲染时跳过）
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    if (r, c) != (min_r, min_c):
                        merge_map[sn][(r, c)] = "skip"

    html_parts = []
    html_parts.append("<!DOCTYPE html>")
    html_parts.append('<html lang="zh-CN">')
    html_parts.append("<head>")
    html_parts.append('<meta charset="UTF-8">')
    html_parts.append(f"<title>{os.path.basename(filepath)} — Excel 预览</title>")
    html_parts.append("<style>")
    html_parts.append("""
body {
    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
    margin: 20px;
    background: #f0f2f5;
    color: #333;
}
h1 {
    font-size: 1.4em;
    margin-bottom: 5px;
}
.file-info {
    color: #888;
    font-size: 0.85em;
    margin-bottom: 20px;
}
.sheet-tabs {
    display: flex;
    gap: 4px;
    margin-bottom: 15px;
    flex-wrap: wrap;
}
.sheet-tab {
    padding: 6px 16px;
    background: #fff;
    border: 1px solid #d0d5dd;
    border-radius: 6px 6px 0 0;
    cursor: pointer;
    font-size: 0.9em;
    transition: 0.15s;
}
.sheet-tab:hover { background: #e8f0fe; }
.sheet-tab.active {
    background: #1a73e8;
    color: #fff;
    border-color: #1a73e8;
}
.sheet-content { display: none; }
.sheet-content.active { display: block; }
.table-wrapper {
    overflow-x: auto;
    border: 1px solid #d0d5dd;
    border-radius: 6px;
    background: #fff;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
}
table {
    border-collapse: collapse;
    white-space: pre-wrap;
    min-width: 100%;
}
td, th {
    padding: 3px 6px;
    min-width: 40px;
    max-width: 300px;
    vertical-align: middle;
}
td:empty::after {
    content: '\\00a0';
}
.col-header {
    background: #f5f6f8 !important;
    color: #666 !important;
    font-weight: bold !important;
    font-size: 0.8em !important;
    text-align: center !important;
    border: 1px solid #d0d5dd !important;
    position: sticky;
    top: 0;
    z-index: 2;
}
.row-header {
    background: #f5f6f8 !important;
    color: #666 !important;
    font-weight: bold !important;
    font-size: 0.8em !important;
    text-align: center !important;
    border: 1px solid #d0d5dd !important;
}
.corner-cell {
    background: #e8eaed !important;
    border: 1px solid #d0d5dd !important;
}
""")
    html_parts.append("</style>")
    html_parts.append("</head>")
    html_parts.append("<body>")

    html_parts.append(f"<h1>📊 {os.path.basename(filepath)}</h1>")
    html_parts.append(f'<div class="file-info">{len(wb.sheetnames)} 个工作表</div>')

    # 标签切换
    tabs = []
    contents = []
    for idx, sheet_name in enumerate(wb.sheetnames):
        active = " active" if idx == 0 else ""
        tabs.append(f'<div class="sheet-tab{active}" onclick="switchSheet(\'{idx}\')">{sheet_name}</div>')

        ws = wb[sheet_name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0

        content = []
        content.append(f'<div class="sheet-content{active}" id="sheet-{idx}">')
        content.append('<div class="table-wrapper"><table>')

        # 列标题行
        content.append('<tr><td class="corner-cell"></td>')
        for c in range(1, cols + 1):
            content.append(f'<td class="col-header">{col_letter(c)}</td>')
        content.append('</tr>')

        # 数据行
        for r in range(1, rows + 1):
            content.append('<tr>')
            # 行号
            content.append(f'<td class="row-header">{r}</td>')

            for c in range(1, cols + 1):
                merge_info = merge_map[sheet_name].get((r, c))
                if merge_info == "skip":
                    continue  # 被合并覆盖的单元格，跳过

                cell = ws.cell(row=r, column=c)
                val = safe_str(cell.value)

                # 收集样式
                css_parts = []
                if cell.font:
                    fc = font_css(cell.font)
                    if fc:
                        css_parts.append(fc)
                if cell.fill:
                    fl = fill_css(cell.fill)
                    if fl:
                        css_parts.append(fl)
                if cell.border:
                    bd = border_css(cell.border)
                    if bd:
                        css_parts.append(bd)
                if cell.alignment:
                    al = align_css(cell.alignment)
                    if al:
                        css_parts.append(al)

                style_attr = f' style="{";".join(css_parts)}"' if css_parts else ""

                # 合并单元格处理
                rowspan = colspan = ""
                if isinstance(merge_info, tuple):
                    rowspan = f' rowspan="{merge_info[0]}"'
                    colspan = f' colspan="{merge_info[1]}"'

                content.append(
                    f'<td{style_attr}{rowspan}{colspan}>{val}</td>'
                )

            content.append('</tr>')

        content.append('</table></div>')
        content.append('</div>')
        contents.append("\n".join(content))

    html_parts.append('<div class="sheet-tabs">')
    html_parts.extend(tabs)
    html_parts.append('</div>')
    html_parts.extend(contents)

    html_parts.append("<script>")
    html_parts.append("""
function switchSheet(idx) {
    document.querySelectorAll('.sheet-tab').forEach((t, i) => t.classList.toggle('active', i == idx));
    document.querySelectorAll('.sheet-content').forEach((c, i) => c.classList.toggle('active', i == idx));
}
""")
    html_parts.append("</script>")
    html_parts.append("</body></html>")

    html_str = "\n".join(html_parts)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_str)

    wb.close()
    print(f"✅ HTML 已导出: {output_path}")
    print(f"   文件大小: {os.path.getsize(output_path):,} 字节")


# ────────────────────────────────────────
#  Diff 命令：对比两个 Excel
# ────────────────────────────────────────

def cmd_diff(file_a, file_b, output_html=None):
    """对比两个 Excel 文件，高亮差异"""
    wb_a = load_workbook(file_a, data_only=True)
    wb_b = load_workbook(file_b, data_only=True)

    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)

    print(f"📄 A: {os.path.basename(file_a)}")
    print(f"📄 B: {os.path.basename(file_b)}")
    print("=" * 70)

    # Sheet 级别差异
    only_a = sheets_a - sheets_b
    only_b = sheets_b - sheets_a
    common = sheets_a & sheets_b

    if only_a:
        print(f"\n🗑 仅在 A 中存在的工作表: {', '.join(sorted(only_a))}")
    if only_b:
        print(f"\n🆕 仅在 B 中存在的工作表: {', '.join(sorted(only_b))}")

    total_diffs = 0
    diff_details = {}  # sheet -> [(cell_ref, val_a, val_b)]

    for sheet_name in sorted(common):
        ws_a = wb_a[sheet_name]
        ws_b = wb_b[sheet_name]

        max_r = max(ws_a.max_row or 0, ws_b.max_row or 0)
        max_c = max(ws_a.max_column or 0, ws_b.max_column or 0)

        sheet_diffs = []
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                va = safe_str(ws_a.cell(row=r, column=c).value) if r <= (ws_a.max_row or 0) and c <= (ws_a.max_column or 0) else ""
                vb = safe_str(ws_b.cell(row=r, column=c).value) if r <= (ws_b.max_row or 0) and c <= (ws_b.max_column or 0) else ""
                if va != vb:
                    sheet_diffs.append((f"{col_letter(c)}{r}", va, vb))

        diff_details[sheet_name] = sheet_diffs
        total_diffs += len(sheet_diffs)

        if sheet_diffs:
            print(f"\n── Sheet: 「{sheet_name}」({len(sheet_diffs)} 处差异)──")
            for ref, va, vb in sheet_diffs[:30]:
                va_short = (va[:30] + "...") if len(va) > 30 else va
                vb_short = (vb[:30] + "...") if len(vb) > 30 else vb
                print(f"  {ref}:  [{va_short}]  →  [{vb_short}]")
            if len(sheet_diffs) > 30:
                print(f"  ... 还有 {len(sheet_diffs) - 30} 处差异")
        else:
            print(f"\n── Sheet: 「{sheet_name}」✅ 完全相同")

    print("\n" + "=" * 70)
    if total_diffs == 0:
        print("✅ 两个文件完全相同!")
    else:
        print(f"🔍 共发现 {total_diffs} 处单元格差异")

    wb_a.close()
    wb_b.close()

    # 如果指定了输出 HTML，生成可视化的 diff HTML
    if output_html:
        _generate_diff_html(file_a, file_b, diff_details, output_html)


def _generate_diff_html(file_a, file_b, diff_details, output_path):
    """生成可视化的 diff HTML 文件，用颜色高亮差异"""
    wb_b = load_workbook(file_b, data_only=True)

    merge_map = {}
    for sn in wb_b.sheetnames:
        ws = wb_b[sn]
        merge_map[sn] = {}
        for merged_range in ws.merged_cells.ranges:
            min_r, min_c = merged_range.min_row, merged_range.min_col
            rowspan = merged_range.max_row - min_r + 1
            colspan = merged_range.max_col - min_c + 1
            merge_map[sn][(min_r, min_c)] = (rowspan, colspan)
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    if (r, c) != (min_r, min_c):
                        merge_map[sn][(r, c)] = "skip"

    # 构建 diff 集合
    diff_set = {}
    for sn, diffs in diff_details.items():
        diff_set[sn] = set()
        for ref, _, _ in diffs:
            diff_set[sn].add(ref)

    html_parts = []
    html_parts.append("<!DOCTYPE html>")
    html_parts.append('<html lang="zh-CN">')
    html_parts.append("<head>")
    html_parts.append('<meta charset="UTF-8">')
    html_parts.append(f"<title>Diff: {os.path.basename(file_a)} vs {os.path.basename(file_b)}</title>")
    html_parts.append("<style>")
    html_parts.append("""
body { font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; margin: 20px; background:#f0f2f5; color:#333; }
h1 { font-size:1.3em; }
.file-info { color:#888; font-size:0.85em; margin-bottom:15px; }
.legend { display:flex; gap:15px; margin-bottom:15px; font-size:0.85em; }
.legend-item { display:flex; align-items:center; gap:6px; }
.legend-swatch { width:18px; height:18px; border-radius:3px; }
.sheet-tabs { display:flex; gap:4px; margin-bottom:15px; flex-wrap:wrap; }
.sheet-tab { padding:6px 16px; background:#fff; border:1px solid #d0d5dd; border-radius:6px 6px 0 0; cursor:pointer; font-size:0.9em; }
.sheet-tab:hover { background:#e8f0fe; }
.sheet-tab.active { background:#1a73e8; color:#fff; border-color:#1a73e8; }
.sheet-content { display:none; }
.sheet-content.active { display:block; }
.table-wrapper { overflow-x:auto; border:1px solid #d0d5dd; border-radius:6px; background:#fff; box-shadow:0 1px 3px rgba(0,0,0,0.08); }
table { border-collapse:collapse; white-space:pre-wrap; min-width:100%; }
td, th { padding:3px 6px; min-width:40px; max-width:300px; vertical-align:middle; }
td:empty::after { content:'\\00a0'; }
.col-header { background:#f5f6f8 !important; color:#666 !important; font-weight:bold !important; font-size:0.8em !important; text-align:center !important; border:1px solid #d0d5dd !important; position:sticky; top:0; z-index:2; }
.row-header { background:#f5f6f8 !important; color:#666 !important; font-weight:bold !important; font-size:0.8em !important; text-align:center !important; border:1px solid #d0d5dd !important; }
.corner-cell { background:#e8eaed !important; border:1px solid #d0d5dd !important; }
.diff-cell { background-color:#fff3cd !important; }
.diff-cell::before { content:'⚠ '; color:#e65100; font-size:0.7em; }
""")
    html_parts.append("</style></head><body>")

    fa = os.path.basename(file_a)
    fb = os.path.basename(file_b)
    html_parts.append(f"<h1>🔍 Diff: {fa} → {fb}</h1>")

    total = sum(len(v) for v in diff_details.values())
    html_parts.append(f'<div class="file-info">{len(wb_b.sheetnames)} 个工作表 · {total} 处单元格差异</div>')
    html_parts.append('<div class="legend">')
    html_parts.append('<div class="legend-item"><div class="legend-swatch" style="background:#fff3cd"></div> 有变化的单元格</div>')
    html_parts.append('</div>')

    # Tabs
    tabs = []
    contents = []
    for idx, sheet_name in enumerate(wb_b.sheetnames):
        active = " active" if idx == 0 else ""
        count = len(diff_details.get(sheet_name, []))
        badge = f" ({count})" if count else ""
        tabs.append(f'<div class="sheet-tab{active}" onclick="switchSheet(\'{idx}\')">{sheet_name}{badge}</div>')

        ws = wb_b[sheet_name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        sheet_diffs = diff_set.get(sheet_name, set())

        content = []
        content.append(f'<div class="sheet-content{active}" id="sheet-{idx}">')
        content.append('<div class="table-wrapper"><table>')

        # 列标题
        content.append('<tr><td class="corner-cell"></td>')
        for c in range(1, cols + 1):
            content.append(f'<td class="col-header">{col_letter(c)}</td>')
        content.append('</tr>')

        for r in range(1, rows + 1):
            content.append('<tr>')
            content.append(f'<td class="row-header">{r}</td>')
            for c in range(1, cols + 1):
                cell_ref = f"{col_letter(c)}{r}"
                merge_info = merge_map[sheet_name].get((r, c))
                if merge_info == "skip":
                    continue

                cell = ws.cell(row=r, column=c)
                val = safe_str(cell.value)
                classes = []
                if cell_ref in sheet_diffs:
                    classes.append("diff-cell")

                css_parts = []
                if cell.font: css_parts.append(font_css(cell.font))
                if cell.fill: css_parts.append(fill_css(cell.fill))
                if cell.border: css_parts.append(border_css(cell.border))
                if cell.alignment: css_parts.append(align_css(cell.alignment))

                class_attr = f' class="{" ".join(classes)}"' if classes else ""
                style_attr = f' style="{";".join(css_parts)}"' if css_parts else ""

                rowspan = colspan = ""
                if isinstance(merge_info, tuple):
                    rowspan = f' rowspan="{merge_info[0]}"'
                    colspan = f' colspan="{merge_info[1]}"'

                content.append(
                    f'<td{class_attr}{style_attr}{rowspan}{colspan}>{val}</td>'
                )
            content.append('</tr>')
        content.append('</table></div></div>')
        contents.append("\n".join(content))

    html_parts.append('<div class="sheet-tabs">')
    html_parts.extend(tabs)
    html_parts.append('</div>')
    html_parts.extend(contents)
    html_parts.append("<script>function switchSheet(i){document.querySelectorAll('.sheet-tab').forEach((t,n)=>t.classList.toggle('active',n==i));document.querySelectorAll('.sheet-content').forEach((c,n)=>c.classList.toggle('active',n==i));}</script>")
    html_parts.append("</body></html>")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html_parts))

    wb_b.close()
    print(f"\n✅ Diff HTML 已导出: {output_path}")
    print(f"   文件大小: {os.path.getsize(output_path):,} 字节")


# ────────────────────────────────────────
#  主入口
# ────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Excel 可视化工具 — dump / html / diff",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python excel_viz.py dump  data.xlsx
  python excel_viz.py html  data.xlsx preview.html
  python excel_viz.py diff  old.xlsx new.xlsx diff.html
        """
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_dump = sub.add_parser("dump", help="快速摸底：打印所有 Sheet 的结构信息")
    p_dump.add_argument("file", help="Excel 文件路径")

    p_html = sub.add_parser("html", help="导出为保留颜色/边框/合并单元格的 HTML")
    p_html.add_argument("file", help="Excel 文件路径")
    p_html.add_argument("output", nargs="?", default=None, help="输出 HTML 路径 (可选)")

    p_diff = sub.add_parser("diff", help="对比两个 Excel 的差异")
    p_diff.add_argument("file_a", help="文件 A")
    p_diff.add_argument("file_b", help="文件 B")
    p_diff.add_argument("output", nargs="?", default=None, help="输出 diff HTML 路径 (可选)")

    args = parser.parse_args()

    if args.cmd == "dump":
        cmd_dump(args.file)
    elif args.cmd == "html":
        cmd_html(args.file, args.output)
    elif args.cmd == "diff":
        cmd_diff(args.file_a, args.file_b, args.output)


if __name__ == "__main__":
    main()
